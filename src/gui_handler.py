# region Imports
import os
import sys
from nicegui import app, ui, events
from queue import Queue
from pydub import AudioSegment
import math
import glob
import openpyxl
import simpleaudio
from typing import List
import src.print_color as pc
from src.logger_setup import gui_logger as logger

# endregion Imports


# region Constants
LANGUAGES = [
    "Auto",
    "Afrikaans",
    "Arabic",
    "Armenian",
    "Azerbaijani",
    "Belarusian",
    "Bosnian",
    "Bulgarian",
    "Catalan",
    "Chinese",
    "Croatian",
    "Czech",
    "Danish",
    "Dutch",
    "English",
    "Estonian",
    "Finnish",
    "French",
    "Galician",
    "German",
    "Greek",
    "Hebrew",
    "Hindi",
    "Hungarian",
    "Icelandic",
    "Indonesian",
    "Italian",
    "Japanese",
    "Kannada",
    "Kazakh",
    "Korean",
    "Latvian",
    "Lithuanian",
    "Macedonian",
    "Malay",
    "Marathi",
    "Maori",
    "Nepali",
    "Norwegian",
    "Persian",
    "Polish",
    "Portuguese",
    "Romanian",
    "Russian",
    "Serbian",
    "Slovak",
    "Slovenian",
    "Spanish",
    "Swahili",
    "Swedish",
    "Tagalog",
    "Tamil",
    "Thai",
    "Turkish",
    "Ukrainian",
    "Urdu",
    "Vietnamese",
    "Welsh",
]
# endregion Constants


# region Utility Functions
def list_downloaded_models(cache_dir="model_cache"):
    """
    List downloaded models from the model_cache directory.
    Ignores hidden directories (starting with '.')
    Returns a list of available model names.
    """
    logger.debug(f"Listing downloaded models from {cache_dir}")
    base_path = os.path.join(os.path.dirname(__file__), "..", cache_dir)
    try:
        models = [
            d
            for d in os.listdir(base_path)
            if os.path.isdir(os.path.join(base_path, d)) and not d.startswith(".")
        ]
        logger.info(f"Found {len(models)} downloaded models")
        return models
    except FileNotFoundError:
        logger.warning(f"Model cache directory not found: {base_path}")
        return []


def mark_downloaded_models(model_name, downloaded_models):
    """Add a marker to downloaded models."""
    cache_name = f"models--{model_name.replace('/', '--')}"
    return f"{'✓' if cache_name in downloaded_models else 'X'} {model_name}"


# endregion Utility Functions


# region View Model
class ViewModel:
    def __init__(self):
        logger.info("Initializing ViewModel")
        self.button_file_content = "choose audio / video files"
        self.selected_files = None
        self.label_progress_content = ""
        self.timestamps_enabled = app.storage.general.get("timestamps_enabled", False)
        self.button_run_enabled = False
        self.button_abort_visible = False
        self.spinner_progress_visibility = False
        self.ui_log = ui.log()
        self.file_count = 0
        self.file_count_old = 0
        self.segment_count = 0
        self.segment_done_count = 0
        self.segment_current_progress = 0
        self.selected_output_formats = []
        self.is_transcribing = False
        self.abort_requested = False
        self.model = None
        self.ui_update_queue = Queue()
        ui.timer(0.1, self.process_ui_updates)

    def process_ui_updates(self):
        while not self.ui_update_queue.empty():
            update = self.ui_update_queue.get()
            logger.debug(f"Processing UI update: {update['type']}")
            if update["type"] == "notify":
                ui.notify(update["message"], type=update["notify_type"])
            elif update["type"] == "log":
                if self.ui_log:
                    self.ui_log.push(update["message"])

    def update_label_progress(self):
        logger.debug(
            f"Updating progress - Files: {self.file_count}, Segments: {self.segment_done_count}/{self.segment_count}"
        )
        if self.file_count <= 0:
            self.spinner_progress_visibility = False
            self.label_progress_content = ""
            self.segment_current_progress = 0
            self.segment_count = 0
            self.segment_done_count = 0
            self.ui_update_queue.put(
                {
                    "type": "notify",
                    "message": "finished transcribing",
                    "notify_type": "positive",
                }
            )
            self.play_sound_effect_finished()
        else:
            if self.file_count_old == 0:
                self.spinner_progress_visibility = True
            info = ""
            if self.file_count == 1:
                info += f"transcribing {self.file_count} file"
            else:
                info += f"transcribing {self.file_count} files"
            if self.segment_count == 1:
                info += (
                    f" ({self.segment_done_count} / {self.segment_count} segment done)"
                )
            else:
                info += (
                    f" ({self.segment_done_count} / {self.segment_count} segments done)"
                )
            if self.segment_current_progress > 0:
                info += f"\ncurrent segment {self.segment_current_progress}% done"
            self.label_progress_content = info
        self.file_count_old = self.file_count

    def update_button_states(self):
        """Update button states based on file and format selection"""
        logger.debug("Updating button states")
        if self.is_transcribing:
            self.button_run_enabled = False
            self.button_abort_visible = True
        else:
            self.button_abort_visible = False
            if self.selected_files is None or len(self.selected_files) == 0:
                self.button_file_content = "choose audio / video files"
                self.button_run_enabled = False
            else:
                self.button_file_content = (
                    "1 file selected"
                    if len(self.selected_files) == 1
                    else f"{len(self.selected_files)} files selected"
                )
                formats = app.storage.general.get("selected_output_format", [])
                self.button_run_enabled = bool(formats) and bool(self.selected_files)

    def update_select_output_formats(self, e: events.ValueChangeEventArguments):
        """Update output formats"""
        logger.info(f"Updating output formats: {e.value}")
        self.selected_output_formats = e.value
        app.storage.general["selected_output_format"] = e.value
        self.update_button_states()

    def update_model_selection(self, e: events.ValueChangeEventArguments):
        """Update selected model"""
        logger.info(f"Updating model selection: {e.value}")
        # Extract model name without the ✓/X prefix
        model_name = e.value.split(" ", 1)[1] if e.value else None
        app.storage.general["selected_model"] = model_name

    def update_language_selection(self, e: events.ValueChangeEventArguments):
        """Update selected language"""
        logger.info(f"Updating language selection: {e.value}")
        app.storage.general["selected_language"] = e.value

    def toggle_timestamps(self, e: events.ValueChangeEventArguments):
        """Toggle timestamps in transcription"""
        logger.info(f"Toggling timestamps: {e.value}")
        self.timestamps_enabled = e.value
        app.storage.general["timestamps_enabled"] = e.value

    async def start_transcription(self):
        """Handle start button click"""
        logger.info("Starting transcription process")
        if not self.is_transcribing:
            self.is_transcribing = True
            self.update_button_states()
            try:
                from src.file_handler import transcribe_files

                # Store the model instance returned from transcribe_files
                result = await transcribe_files(
                    self.selected_files,
                    app.storage.general["selected_output_format"],
                    self,
                    self.model,
                    self.update_ui,
                    model_name=app.storage.general["selected_model"],
                    language=app.storage.general["selected_language"],
                    timestamps_enabled=self.timestamps_enabled,
                )
                # Update the model instance if a new one was created
                if isinstance(result, tuple) and len(result) == 2:
                    self.model = result[0]
            except Exception as e:
                self.update_ui(f"Transcription error: {str(e)}", "negative")
            finally:
                self.is_transcribing = False
                self.abort_requested = False
                self.file_count = 0
                self.update_label_progress()
                self.update_button_states()

    def update_ui(self, message, notify_type=None):
        """Queue UI updates to be processed in the main thread"""
        logger.debug(f"UI Update: {message} (type: {notify_type})")
        if notify_type:
            self.ui_update_queue.put(
                {"type": "notify", "message": message, "notify_type": notify_type}
            )
        self.ui_update_queue.put({"type": "log", "message": message})

    def abort_transcription(self):
        """Handle abort button click"""
        logger.info("Transcription abort requested")
        self.abort_requested = True
        self.ui_update_queue.put(
            {
                "type": "notify",
                "message": "Aborting transcription...",
                "notify_type": "warning",
            }
        )

    def toggle_mute(self):
        logger.debug("Toggling mute state")
        app.storage.general["mute"] = not app.storage.general["mute"]
        self.ui_update_queue.put(
            {
                "type": "notify",
                "message": "Sound "
                + ("unmuted" if not app.storage.general["mute"] else "muted"),
                "notify_type": "info",
            }
        )

    def toggle_dark_mode(self):
        logger.debug("Toggling dark mode")
        app.storage.general["dark"] = not app.storage.general.get("dark", False)
        # ui.dark_mode is safe to call directly as it's a global UI state change
        ui.dark_mode(app.storage.general["dark"])
        self.ui_update_queue.put(
            {
                "type": "notify",
                "message": "Dark mode "
                + ("enabled" if app.storage.general["dark"] else "disabled"),
                "notify_type": "info",
            }
        )

    def play_sound_effect_finished(self):
        logger.debug("Playing finished sound effect")
        if not app.storage.general.get("mute", False):
            sound_effect_path = "sound_effect_finished.wav"
            if not os.path.isfile(sound_effect_path):
                sound_effect_path = os.path.join(
                    "_internal", "sound_effect_finished.wav"
                )
            if sound_effect_path is not None:
                wave_obj = simpleaudio.WaveObject.from_wave_file(sound_effect_path)
                wave_obj.play()


# endregion View Model


# region Audio Processing
class AudioSplitter:
    @staticmethod
    def split_audio(
        file: str, max_size: int = 25000000, overlap_ms: int = 2000
    ) -> List[tuple]:
        logger.info(f"Splitting audio file: {file}")
        split_files = []
        segment_count = AudioSplitter.get_segment_count(file, max_size)
        if segment_count > 1:
            temp_dir = AudioSplitter.create_temp_dir()
            song = AudioSegment.from_file(file)
            pc.print_info(f"\nsplitting {file} into {segment_count} parts")
            segment_length_ms = len(song) / segment_count
            overlap = overlap_ms if segment_length_ms >= overlap_ms else 0

            for i in range(segment_count):
                seg_start_ms = (
                    math.floor(i * segment_length_ms - overlap / 2) if i > 0 else 0
                )
                seg_end_ms = (
                    math.ceil((i + 1) * segment_length_ms + overlap / 2)
                    if i < segment_count - 1
                    else None
                )

                segment = song[seg_start_ms:seg_end_ms]
                segment_filename = os.path.join(
                    temp_dir,
                    f"{os.path.splitext(os.path.basename(file))[0]}_segment{i}{os.path.splitext(file)[1]}",
                )
                segment.export(segment_filename)
                pc.print_info(f"\n  > saved segment {segment_filename}")
                split_files.append((segment_filename, seg_start_ms))
            pc.print_info(f"\nfinished splitting {file}")
        else:
            split_files.append((file, 0))
        return split_files

    @staticmethod
    def get_segment_count(file: str, max_size: int = 25000000) -> int:
        logger.debug(f"Calculating segments for file: {file}")
        file_size = os.path.getsize(file)
        return math.ceil(file_size / max_size) if file_size > max_size else 1

    @staticmethod
    def get_temp_dir() -> str:
        logger.debug("Getting temp directory path")
        if getattr(sys, "frozen", False):
            return os.path.join(os.path.dirname(sys.executable), ".temp")
        elif __file__:
            return os.path.join(os.path.dirname(__file__), ".temp")
        return os.path.join(os.getcwd(), ".temp")

    @staticmethod
    def create_temp_dir() -> str:
        logger.info("Creating temporary directory")
        temp_dir = AudioSplitter.get_temp_dir()
        pc.print_info(f"\ntemp dir: {temp_dir}")
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        return temp_dir

    @staticmethod
    def clear_temp_dir():
        logger.info("Clearing temporary directory")
        temp_dir = AudioSplitter.get_temp_dir()
        files = glob.glob(os.path.join(temp_dir, "*"))
        for f in files:
            os.remove(f)


# endregion Audio Processing


# region Transcription Handler
class TranscriptionHandler:
    @staticmethod
    def save_result(result: dict, output_formats: List[str], file_path: str):
        logger.info(f"Saving transcription results for {file_path}")
        output_dir = os.path.dirname(os.path.realpath(file_path))

        for ext in output_formats:
            output_filename = f"{os.path.splitext(file_path)[0]}.{ext}"
            logger.debug(f"Saving {ext} format to: {output_filename}")

            if ext == "xlsx":
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = "start"
                sheet.cell(row=1, column=2).value = "end"
                sheet.cell(row=1, column=3).value = "text"

                for i, segment in enumerate(result["segments"], 2):
                    sheet.cell(row=i, column=1).value = segment["start"]
                    sheet.cell(row=i, column=2).value = segment["end"]
                    sheet.cell(row=i, column=3).value = segment["text"]

                wb.save(os.path.join(output_dir, output_filename))
            else:
                # Handle other formats (txt, srt, vtt, etc.)
                with open(
                    os.path.join(output_dir, output_filename), "w", encoding="utf-8"
                ) as f:
                    if ext == "txt":
                        f.write(result["text"])
                    elif ext in ["srt", "vtt"]:
                        for i, segment in enumerate(result["segments"], 1):
                            f.write(f"{i}\n")
                            start = TranscriptionHandler.format_timestamp(
                                segment["start"]
                            )
                            end = TranscriptionHandler.format_timestamp(segment["end"])
                            f.write(f"{start} --> {end}\n")
                            f.write(f"{segment['text']}\n\n")

    @staticmethod
    def format_timestamp(seconds: float) -> str:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        msecs = int((seconds % 1) * 1000)
        return f"{hours:02d}:{minutes:02d}:{secs:02d},{msecs:03d}"


# endregion Transcription Handler
